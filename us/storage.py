"""
storage.py — Simpan hasil ke file xlsx harian

Total Score di xlsx dihitung ulang dengan weight terbaru saat save,
sehingga jika weight diupdate via ML, xlsx langsung reflect weight baru.
"""

import os
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from config import OUTPUT_DIR
from weight_manager import load_weights, apply_weights, FEATURES

logger = logging.getLogger(__name__)

HEADERS = [
    "Ticker", "Date", "Price", "Change%",
    "VSA", "FSA", "VFA", "WCC", "SRST", "RSI", "MACD", "MA",
    "IP Raw", "IP Score", "Tight", "Total Score",
]

COLOR_HEADER = "1F4E79"
COLOR_GREEN  = "C6EFCE"
COLOR_RED    = "FFC7CE"
COLOR_WHITE  = "FFFFFF"


def _recalc_total(r: dict) -> float:
    """
    Hitung ulang total score dengan weight terbaru untuk ticker ini.
    Dipakai saat save_to_xlsx agar xlsx selalu reflect weight terkini.
    """
    ticker  = r.get("ticker", "")
    weights = load_weights(ticker)
    scores  = {
        "vsa":      r.get("vsa",      0),
        "fsa":      r.get("fsa",      0),
        "vfa":      r.get("vfa",      0),
        "wcc":      r.get("wcc",      0),
        "srst":     r.get("srst",     0),
        "rsi":      r.get("rsi",      0),
        "macd":     r.get("macd",     0),
        "ma":       r.get("ma",       0),
        "ip_score": r.get("ip_score", 0),
        "tight":    r.get("tight",    0),
    }
    return round(apply_weights(scores, weights), 4)


def _row_from_result(r: dict) -> list:
    total = _recalc_total(r)
    return [
        r["ticker"], r["date"], r["price"], r["change"],
        r["vsa"], r.get("fsa", 0), r.get("vfa", 0), r.get("wcc", 0), r.get("srst", 0),
        r["rsi"], r["macd"], r["ma"],
        r["ip_raw"], r["ip_score"], r.get("tight", 0), total,
    ]


def save_to_xlsx(results: list[dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    today    = datetime.today().strftime("%Y-%m-%d")
    filepath = os.path.join(OUTPUT_DIR, f"score_{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Scores"

    hdr_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    hdr_font = Font(color=COLOR_WHITE, bold=True)
    center   = Alignment(horizontal="center")

    for col, h in enumerate(HEADERS, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center

    green_fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
    red_fill   = PatternFill(start_color=COLOR_RED,   end_color=COLOR_RED,   fill_type="solid")

    for row_idx, r in enumerate(results, 2):
        row_data = _row_from_result(r)
        total    = row_data[-1]   # index 15 = Total Score (sudah recalc)

        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center

        if total > 4:
            fill = green_fill
        elif total < -4:
            fill = red_fill
        else:
            fill = None

        if fill:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filepath)
    logger.info(f"Hasil disimpan: {filepath}")
    return filepath


def update_xlsx_weights(results: list[dict]) -> str:
    """
    Buka xlsx hari ini dan update kolom Total Score dengan weight terbaru.
    Dipanggil setelah ML selesai update weight agar xlsx langsung ter-refresh.

    Args:
        results: list dict hasil scoring (dari /9 terakhir)

    Returns:
        filepath xlsx yang diupdate
    """
    today    = datetime.today().strftime("%Y-%m-%d")
    filepath = os.path.join(OUTPUT_DIR, f"score_{today}.xlsx")

    if not os.path.exists(filepath):
        logger.warning(f"Tidak ada xlsx hari ini: {filepath}")
        return filepath

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Cari index kolom Total Score dan Ticker
        header_row = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
        col_total  = header_row.get("Total Score")
        col_ticker = header_row.get("Ticker")

        if not col_total or not col_ticker:
            logger.warning("Kolom Total Score atau Ticker tidak ditemukan di xlsx")
            return filepath

        green_fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
        red_fill   = PatternFill(start_color=COLOR_RED,   end_color=COLOR_RED,   fill_type="solid")
        white_fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

        # Build lookup dari results
        result_map = {r["ticker"]: r for r in results}

        for row in range(2, ws.max_row + 1):
            ticker_cell = ws.cell(row, col_ticker)
            ticker = str(ticker_cell.value or "").strip().upper()
            if not ticker or ticker not in result_map:
                continue

            r     = result_map[ticker]
            total = _recalc_total(r)

            ws.cell(row, col_total).value = total

            # Update warna baris
            if total > 4:
                fill = green_fill
            elif total < -4:
                fill = red_fill
            else:
                fill = white_fill

            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = fill

        wb.save(filepath)
        logger.info(f"[xlsx] Updated dengan weight terbaru: {filepath}")

    except Exception as e:
        logger.error(f"Gagal update xlsx: {e}")

    return filepath
