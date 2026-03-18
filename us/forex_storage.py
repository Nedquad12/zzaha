"""
forex_storage.py — Simpan hasil scoring forex ke file xlsx harian

File output: /home/ec2-user/us/forex_score_{YYYY-MM-DD}.xlsx
Dipisah dari saham agar tidak saling menimpa.
"""

import os
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from config import OUTPUT_DIR
from forex_scorer import load_forex_weights, apply_forex_weights, FEATURES

logger = logging.getLogger(__name__)

HEADERS = [
    "Pair", "Date", "Price", "Change%",
    "VSA", "FSA", "VFA", "WCC", "SRST", "RSI", "MACD", "MA",
    "IP Raw", "IP Score", "Tight", "Total Score",
]

COLOR_HEADER = "1F4E79"
COLOR_GREEN  = "C6EFCE"
COLOR_RED    = "FFC7CE"
COLOR_WHITE  = "FFFFFF"


def _recalc_total(r: dict) -> float:
    """Hitung ulang total score dengan weight terbaru."""
    scores = {
        "vsa":      r.get("vsa", 0),
        "fsa":      r.get("fsa", 0),
        "vfa":      r.get("vfa", 0),
        "wcc":      r.get("wcc", 0),
        "srst":     r.get("srst", 0),
        "rsi":      r.get("rsi", 0),
        "macd":     r.get("macd", 0),
        "ma":       r.get("ma", 0),
        "ip_score": r.get("ip_score", 0),
        "tight":    r.get("tight", 0),
    }
    weights = load_forex_weights(r.get("pair", r.get("ticker", "")))
    return apply_forex_weights(scores, weights)


def save_forex_to_xlsx(results: list[dict]) -> str:
    """
    Simpan hasil scoring forex ke xlsx.
    Return path file yang disimpan.
    """
    today    = datetime.today().strftime("%Y-%m-%d")
    filename = f"forex_score_{today}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forex Scores"

    # ── Header ────────────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Data ─────────────────────────────────────────────────────────────────
    green_fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
    red_fill   = PatternFill(start_color=COLOR_RED,   end_color=COLOR_RED,   fill_type="solid")
    white_fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

    sorted_results = sorted(results, key=lambda x: x["total"], reverse=True)

    for row_idx, r in enumerate(sorted_results, 2):
        total = _recalc_total(r)
        pair  = r.get("pair", r.get("ticker", ""))

        values = [
            pair,
            r.get("date", today),
            r.get("price", 0),
            r.get("change", 0),
            r.get("vsa", 0),
            r.get("fsa", 0),
            r.get("vfa", 0),
            r.get("wcc", 0),
            r.get("srst", 0),
            r.get("rsi", 0),
            r.get("macd", 0),
            r.get("ma", 0),
            r.get("ip_raw", 0),
            r.get("ip_score", 0),
            r.get("tight", 0),
            round(total, 4),
        ]

        fill = green_fill if total > 4 else (red_fill if total < -4 else white_fill)

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # ── Auto column width ─────────────────────────────────────────────────────
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    wb.save(filepath)
    logger.info(f"[Forex] xlsx disimpan: {filepath}")
    return filepath
