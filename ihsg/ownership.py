"""
ownership.py — Handler data kepemilikan untuk /ex dan /xlsx

Sumber data : /home/ec2-user/database/data  (Excel bulanan)
Kolom kunci : Date, Code, Type, Price,
              Local IS/CP/PF/IB/ID/MF/SC/FD/OT,
              Foreign IS/CP/PF/IB/ID/MF/SC/FD/OT

/ex BBCA   → 2 gambar PNG: pie donut Local + pie donut Foreign (data terbaru)
/xlsx BBCA → 1 file Excel (Shares + Value + Ownership %, sama persis lama)
"""

import gc
import glob
import io
import logging
import os

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Konstanta ─────────────────────────────────────────────────────────────────
DATA_FOLDER = "/home/ec2-user/database/data"

LOCAL_CATS = [
    'Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID',
    'Local MF', 'Local SC', 'Local FD', 'Local OT',
]
FOREIGN_CATS = [
    'Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID',
    'Foreign MF', 'Foreign SC', 'Foreign FD', 'Foreign OT',
]

LABELS = {
    'IS': 'Asuransi',  'CP': 'Korporat',     'PF': 'Dana Pensiun',
    'IB': 'Bank',      'ID': 'Ritel',         'MF': 'Reksadana',
    'SC': 'Sekuritas', 'FD': 'Foundation',    'OT': 'Lainnya',
}

LOCAL_COLORS = [
    '#1f77b4', '#2ca02c', '#ff7f0e', '#9467bd', '#8c564b',
    '#e377c2', '#17becf', '#bcbd22', '#d62728',
]
FOREIGN_COLORS = [
    '#d62728', '#ff7f0e', '#e377c2', '#9467bd', '#17becf',
    '#2ca02c', '#8c564b', '#bcbd22', '#1f77b4',
]

WATERMARK = 'Membahas Saham Indonesia'


# ══════════════════════════════════════════════════════════════════════════════
#  Data loader
# ══════════════════════════════════════════════════════════════════════════════

def _load_all() -> pd.DataFrame | None:
    """Gabungkan semua Excel di DATA_FOLDER → satu DataFrame."""
    if not os.path.exists(DATA_FOLDER):
        logger.error(f"Folder tidak ditemukan: {DATA_FOLDER}")
        return None

    files = []
    for ext in ('*.xlsx', '*.xls', '*.XLSX', '*.XLS'):
        files.extend(glob.glob(os.path.join(DATA_FOLDER, ext)))

    if not files:
        logger.warning("Tidak ada file Excel di DATA_FOLDER.")
        return None

    dfs = []
    for fp in sorted(files):
        try:
            dfs.append(pd.read_excel(fp))
            logger.info(f"Loaded {os.path.basename(fp)}")
        except Exception as e:
            logger.error(f"Gagal baca {fp}: {e}")

    if not dfs:
        return None

    df = pd.concat(dfs, ignore_index=True)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.dropna(subset=['Date'])
    df = df.drop_duplicates(subset=['Date', 'Code'], keep='last')
    df = df.sort_values('Date', ascending=True).reset_index(drop=True)

    for col in LOCAL_CATS + FOREIGN_CATS + ['Price']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    return df


def _get_stock(code: str) -> pd.DataFrame | None:
    """Semua baris untuk satu kode saham, ascending by Date."""
    df = _load_all()
    if df is None or 'Code' not in df.columns:
        return None
    result = df[df['Code'].str.upper() == code.upper()].copy()
    return result if not result.empty else None


# ══════════════════════════════════════════════════════════════════════════════
#  /ex — 2 pie chart PNG
# ══════════════════════════════════════════════════════════════════════════════

def _make_pie_image(
    row: pd.Series,
    cats: list,
    colors: list,
    title: str,
    code: str,
    date_str: str,
    price_str: str,
) -> io.BytesIO:
    """Buat 1 gambar donut dari 1 baris data terbaru."""

    entries = []
    for cat, color in zip(cats, colors):
        val = float(row.get(cat, 0))
        if val > 0:
            key   = cat.split()[-1]
            label = LABELS.get(key, key)
            entries.append((val, label, color))

    fig, ax = plt.subplots(figsize=(9, 7))
    fig.patch.set_facecolor('white')

    if not entries:
        ax.text(0.5, 0.5, 'Tidak ada data', ha='center', va='center',
                fontsize=14, transform=ax.transAxes)
        ax.axis('off')
    else:
        sizes  = [e[0] for e in entries]
        labels = [e[1] for e in entries]
        clrs   = [e[2] for e in entries]
        total  = sum(sizes)
        pct    = [s / total * 100 for s in sizes]
        legend = [f"{l}  {p:.1f}%" for l, p in zip(labels, pct)]

        wedges, _ = ax.pie(
            sizes, colors=clrs,
            startangle=90, counterclock=False,
            wedgeprops=dict(width=0.55, edgecolor='white', linewidth=1.5),
        )
        ax.legend(
            wedges, legend,
            title="Alokasi",
            loc="center left", bbox_to_anchor=(1.02, 0.5),
            fontsize=10, title_fontsize=11,
        )
        ax.axis('equal')

    ax.set_title(title, fontsize=14, fontweight='bold', pad=14)
    fig.suptitle(
        f"{code.upper()}   {date_str}   {price_str}",
        fontsize=11, color='#555555', y=0.02,
    )

    # Watermark
    fig.text(0.5, 0.5, WATERMARK, fontsize=40, color='gray',
             ha='center', va='center', alpha=0.12, rotation=30,
             transform=fig.transFigure, zorder=0)

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=180, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    plt.close(fig)
    gc.collect()
    return buf


def create_ownership_charts(code: str) -> tuple[io.BytesIO | None, io.BytesIO | None]:
    """
    Return (buf_local, buf_foreign).
    Keduanya None jika data tidak ditemukan.
    """
    df = _get_stock(code)
    if df is None:
        return None, None

    latest    = df.iloc[-1]
    date_str  = latest['Date'].strftime('%d %b %Y')
    price_val = latest.get('Price', 0)
    price_str = f"Rp {int(price_val):,}" if price_val else ''

    buf_local = _make_pie_image(
        row=latest, cats=LOCAL_CATS, colors=LOCAL_COLORS,
        title="🇮🇩  Kepemilikan LOKAL",
        code=code, date_str=date_str, price_str=price_str,
    )
    buf_foreign = _make_pie_image(
        row=latest, cats=FOREIGN_CATS, colors=FOREIGN_COLORS,
        title="🌏  Kepemilikan ASING",
        code=code, date_str=date_str, price_str=price_str,
    )

    return buf_local, buf_foreign


# ══════════════════════════════════════════════════════════════════════════════
#  /forc — 2 bar chart perubahan bulan terbaru vs sebelumnya
# ══════════════════════════════════════════════════════════════════════════════

def _make_change_image(
    row_new: pd.Series,
    row_old: pd.Series,
    cats: list,
    side: str,
    code: str,
    date_new: str,
    date_old: str,
) -> io.BytesIO:
    """Bar chart perubahan (absolut + %) per kategori investor."""

    short    = [LABELS.get(c.split()[-1], c.split()[-1]) for c in cats]
    deltas   = [float(row_new.get(c, 0)) - float(row_old.get(c, 0)) for c in cats]
    old_vals = [float(row_old.get(c, 0)) for c in cats]

    pcts = []
    for d, o in zip(deltas, old_vals):
        pcts.append(d / o * 100 if o != 0 else 0.0)

    colors = ['#2ecc71' if d >= 0 else '#e74c3c' for d in deltas]

    fig, (ax_abs, ax_pct) = plt.subplots(1, 2, figsize=(14, 6))
    fig.patch.set_facecolor('white')
    fig.suptitle(
        f"{'🇮🇩' if side == 'Lokal' else '🌏'}  Perubahan Kepemilikan {side} — {code.upper()}\n"
        f"{date_old}  →  {date_new}",
        fontsize=13, fontweight='bold', y=1.01,
    )

    x = range(len(short))

    # Kiri: absolut
    bars = ax_abs.bar(x, deltas, color=colors, edgecolor='white', linewidth=0.8)
    ax_abs.axhline(0, color='#555555', linewidth=0.8, linestyle='--')
    ax_abs.set_xticks(list(x))
    ax_abs.set_xticklabels(short, rotation=30, ha='right', fontsize=9)
    ax_abs.set_title('Perubahan Absolut (lembar)', fontsize=11, fontweight='bold')
    ax_abs.set_ylabel('Lembar Saham')
    ax_abs.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    for bar, val in zip(bars, deltas):
        if val == 0:
            continue
        va  = 'bottom' if val >= 0 else 'top'
        off = max(abs(val) * 0.02, 1)
        ax_abs.text(
            bar.get_x() + bar.get_width() / 2,
            val + (off if val >= 0 else -off),
            f"{val:+,.0f}", ha='center', va=va, fontsize=7.5, fontweight='bold',
        )
    ax_abs.grid(axis='y', alpha=0.3)

    # Kanan: %
    bars2 = ax_pct.bar(x, pcts, color=colors, edgecolor='white', linewidth=0.8)
    ax_pct.axhline(0, color='#555555', linewidth=0.8, linestyle='--')
    ax_pct.set_xticks(list(x))
    ax_pct.set_xticklabels(short, rotation=30, ha='right', fontsize=9)
    ax_pct.set_title('Perubahan (%)', fontsize=11, fontweight='bold')
    ax_pct.set_ylabel('Persentase (%)')
    ax_pct.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:+.1f}%"))
    for bar, val in zip(bars2, pcts):
        if val == 0:
            continue
        va  = 'bottom' if val >= 0 else 'top'
        off = max(abs(val) * 0.02, 0.1)
        ax_pct.text(
            bar.get_x() + bar.get_width() / 2,
            val + (off if val >= 0 else -off),
            f"{val:+.1f}%", ha='center', va=va, fontsize=7.5, fontweight='bold',
        )
    ax_pct.grid(axis='y', alpha=0.3)

    fig.text(0.5, 0.5, WATERMARK, fontsize=40, color='gray',
             ha='center', va='center', alpha=0.10, rotation=30,
             transform=fig.transFigure, zorder=0)

    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=180, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    plt.close(fig)
    gc.collect()
    return buf


def create_flow_charts(
    code: str, n: int = 1
) -> tuple[io.BytesIO | None, io.BytesIO | None]:
    """
    Return (buf_local, buf_foreign) bar chart perubahan.
    n = berapa bulan ke belakang sebagai pembanding (1-5).
    Misal n=3: bandingkan bulan terbaru vs 3 bulan sebelumnya.
    """
    df = _get_stock(code)
    n  = max(1, min(5, n))          # clamp 1–5
    if df is None or len(df) < n + 1:
        return None, None

    row_new  = df.iloc[-1]
    row_old  = df.iloc[-(n + 1)]
    date_new = row_new['Date'].strftime('%b %Y')
    date_old = row_old['Date'].strftime('%b %Y')

    buf_local = _make_change_image(
        row_new=row_new, row_old=row_old,
        cats=LOCAL_CATS, side='Lokal',
        code=code, date_new=date_new, date_old=date_old,
    )
    buf_foreign = _make_change_image(
        row_new=row_new, row_old=row_old,
        cats=FOREIGN_CATS, side='Asing',
        code=code, date_new=date_new, date_old=date_old,
    )

    return buf_local, buf_foreign


# ══════════════════════════════════════════════════════════════════════════════
#  /top — ranking perubahan kepemilikan semua saham
# ══════════════════════════════════════════════════════════════════════════════

def get_top_changes(
    side: str, category: str, top_n: int = 20
) -> tuple[list[dict], list[dict]] | None:
    """
    Hitung % change bulan terbaru vs sebelumnya untuk semua saham.

    side     : 'local' atau 'foreign'
    category : kode kategori, misal 'IS', 'CP', 'ALL' (total)
    top_n    : jumlah saham per list (default 20)

    Return (top_up, top_down) — masing-masing list 20 saham,
    atau None jika data tidak cukup.
    """
    df = _load_all()
    if df is None or 'Code' not in df.columns:
        return None

    prefix = 'Local' if side == 'local' else 'Foreign'
    if category == 'ALL':
        cats = LOCAL_CATS if side == 'local' else FOREIGN_CATS
    else:
        col = f"{prefix} {category}"
        if col not in df.columns:
            return None
        cats = [col]

    results = []
    for ticker, grp in df.groupby('Code'):
        grp = grp.sort_values('Date')
        if len(grp) < 2:
            continue
        row_new = grp.iloc[-1]
        row_old = grp.iloc[-2]
        val_new = sum(float(row_new.get(c, 0)) for c in cats)
        val_old = sum(float(row_old.get(c, 0)) for c in cats)
        if val_old == 0:
            continue
        pct = (val_new - val_old) / val_old * 100
        results.append({
            'ticker':     str(ticker),
            'old':        val_old,
            'new':        val_new,
            'change_pct': pct,
        })

    if not results:
        return None

    # Top 20 naik (terbesar ke terkecil)
    top_up = sorted(results, key=lambda x: x['change_pct'], reverse=True)[:top_n]
    for i, r in enumerate(top_up, 1):
        r['rank'] = i

    # Top 20 turun (terkecil ke terbesar)
    top_down = sorted(results, key=lambda x: x['change_pct'])[:top_n]
    for i, r in enumerate(top_down, 1):
        r['rank'] = i

    return top_up, top_down


# ══════════════════════════════════════════════════════════════════════════════
#  /xlsx — Excel (sama persis create_excel_report lama)
# ══════════════════════════════════════════════════════════════════════════════

def create_ownership_excel(code: str) -> io.BytesIO | None:
    """Return BytesIO berisi .xlsx, atau None jika data tidak ada."""
    df = _get_stock(code)
    if df is None:
        return None

    data         = df.copy()
    latest_price = float(
        data.sort_values('Date', ascending=False).iloc[0].get('Price', 0)
    )

    wb = Workbook()

    # Style
    header_font       = Font(bold=True, color="FFFFFF")
    header_fill       = PatternFill(start_color="366092", end_color="366092",
                                    fill_type="solid")
    border            = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    center_align      = Alignment(horizontal='center')
    number_format     = '#,##0'
    currency_format   = '_("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"_);_(@_)'
    percentage_format = '0.00"%"'

    # ── Sheet 1: Shares ────────────────────────────────────────────────────
    ws_shares       = wb.active
    ws_shares.title = "Shares"

    headers_shares = (
        ['Date', 'Code', 'Type', 'Price']
        + LOCAL_CATS + ['Total Local']
        + FOREIGN_CATS + ['Total Foreign']
    )
    for col, h in enumerate(headers_shares, 1):
        cell = ws_shares.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.border, cell.alignment = (
            header_font, header_fill, border, center_align
        )

    for ri, (_, r) in enumerate(data.iterrows(), 2):
        ws_shares.cell(ri, 1, r['Date'].strftime('%d-%b-%Y')).border = border
        ws_shares.cell(ri, 2, r.get('Code', '')).border              = border
        ws_shares.cell(ri, 3, r.get('Type', '')).border              = border
        ws_shares.cell(ri, 4, r.get('Price', 0)).border              = border

        for ci, cat in enumerate(LOCAL_CATS, 5):
            c = ws_shares.cell(ri, ci, r.get(cat, 0))
            c.border = border; c.number_format = number_format

        tl = ws_shares.cell(ri, 14, f"=SUM(E{ri}:M{ri})")
        tl.border = border; tl.number_format = number_format

        for ci, cat in enumerate(FOREIGN_CATS, 15):
            c = ws_shares.cell(ri, ci, r.get(cat, 0))
            c.border = border; c.number_format = number_format

        tf = ws_shares.cell(ri, 24, f"=SUM(O{ri}:W{ri})")
        tf.border = border; tf.number_format = number_format

    for col in ws_shares.columns:
        mx = max((len(str(c.value)) for c in col if c.value), default=0)
        ws_shares.column_dimensions[
            get_column_letter(col[0].column)
        ].width = min(mx + 2, 20)

    # ── Sheet 2: Value ─────────────────────────────────────────────────────
    ws_value        = wb.create_sheet(title="Value")

    headers_value = (
        ['Date', 'Code', 'Type', '']
        + LOCAL_CATS + ['Total Local']
        + FOREIGN_CATS + ['Total Foreign']
    )
    for col, h in enumerate(headers_value, 1):
        if h:
            cell = ws_value.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.border, cell.alignment = (
                header_font, header_fill, border, center_align
            )

    for ri, (_, r) in enumerate(data.iterrows(), 2):
        ws_value.cell(ri, 1, r['Date'].strftime('%d-%b-%Y')).border = border
        ws_value.cell(ri, 2, r.get('Code', '')).border              = border
        ws_value.cell(ri, 3, r.get('Type', '')).border              = border
        ws_value.cell(ri, 4, '').border                             = border

        for ci, cat in enumerate(LOCAL_CATS, 5):
            c = ws_value.cell(ri, ci, r.get(cat, 0) * latest_price)
            c.border = border; c.number_format = currency_format

        tl = ws_value.cell(ri, 14, f"=SUM(E{ri}:M{ri})")
        tl.border = border; tl.number_format = currency_format

        for ci, cat in enumerate(FOREIGN_CATS, 15):
            c = ws_value.cell(ri, ci, r.get(cat, 0) * latest_price)
            c.border = border; c.number_format = currency_format

        tf = ws_value.cell(ri, 24, f"=SUM(O{ri}:W{ri})")
        tf.border = border; tf.number_format = currency_format

    # Ownership % di bawah data Value
    last_data_row = len(data) + 1
    pct_start     = last_data_row + 3

    ws_value.cell(pct_start, 1,
                  "OWNERSHIP PERCENTAGE (%)").font = Font(bold=True, size=12)

    ws_value.cell(pct_start + 1, 1, "Date").fill   = header_fill
    ws_value.cell(pct_start + 1, 1).font            = header_font
    ws_value.cell(pct_start + 1, 1).border          = border

    pct_hdrs = (
        ['Local IS%','Local CP%','Local PF%','Local IB%','Local ID%',
         'Local MF%','Local SC%','Local FD%','Local OT%']
        + ['Foreign IS%','Foreign CP%','Foreign PF%','Foreign IB%','Foreign ID%',
           'Foreign MF%','Foreign SC%','Foreign FD%','Foreign OT%']
    )
    for ci, h in enumerate(pct_hdrs, 2):
        cell = ws_value.cell(pct_start + 1, ci, h)
        cell.fill = header_fill; cell.font = header_font; cell.border = border

    for di, (_, r) in enumerate(data.iterrows(), 2):
        prow = pct_start + 1 + (di - 1)
        ws_value.cell(prow, 1, r['Date'].strftime('%d-%b-%Y')).border = border

        for i in range(9):
            cl  = get_column_letter(5 + i)
            c   = ws_value.cell(prow, 2 + i,
                                f"=IF(N{di}=0,0,{cl}{di}/N{di}*100)")
            c.border = border; c.number_format = percentage_format

        for i in range(9):
            cl  = get_column_letter(15 + i)
            c   = ws_value.cell(prow, 11 + i,
                                f"=IF(X{di}=0,0,{cl}{di}/X{di}*100)")
            c.border = border; c.number_format = percentage_format

    for col in ws_value.columns:
        mx = max((len(str(c.value)) for c in col if c.value), default=0)
        ws_value.column_dimensions[
            get_column_letter(col[0].column)
        ].width = min(mx + 2, 20)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
